# format_pcw_json.py

import openpyxl

def fill_template_from_json(gpt_data, template_path, output_path):
    wb = openpyxl.load_workbook(template_path)
    ws = wb['Rev 0']

    # === PARTS LIST ===
    parts = gpt_data.get("parts_list", [])
    start_row_parts = 10
    for idx, part in enumerate(parts):
        ws.cell(row=start_row_parts + idx, column=1, value=part.get("Description", ""))
        ws.cell(row=start_row_parts + idx, column=2, value=part.get("Manufacturer", ""))
        ws.cell(row=start_row_parts + idx, column=3, value=part.get("Part Number", ""))
        ws.cell(row=start_row_parts + idx, column=4, value=part.get("Quantity", 0))
        ws.cell(row=start_row_parts + idx, column=5, value=part.get("Cost/ea", 0))

    ws.cell(row=39, column=6, value=f'=SUM(F{start_row_parts}:F38)')
    ws.cell(row=39, column=8, value=f'=SUM(H{start_row_parts}:H38)')

    # === ENGINEERING LABOR ===
    eng_rows = gpt_data.get("engineering_labor", [])
    for row_idx in range(43, 75):
        task_cell = ws.cell(row=row_idx, column=1).value
        if not task_cell:
            continue
        task_clean = task_cell.strip()
        match = next((r for r in eng_rows if r["Task"].strip() == task_clean), None)
        if match:
            ws.cell(row=row_idx, column=2, value=match.get("# of DWG / Tasks", 0))
            ws.cell(row=row_idx, column=3, value=match.get("Hrs per DWG/Task", 0))
            ws.cell(row=row_idx, column=4, value=match.get("Total Hrs", 0))
        else:
            ws.cell(row=row_idx, column=2, value=0)
            ws.cell(row=row_idx, column=3, value=0)
            ws.cell(row=row_idx, column=4, value=0)

    # === ASSEMBLY + IN-HOUSE TESTING LABOR ===
    assembly = gpt_data.get("assembly_labor", [])
    testing = gpt_data.get("in_house_testing_labor", [])

    def find_task(task_name, source):
        return next((t for t in source if t["Task"].strip() == task_name), None)

    for row in ws.iter_rows(min_row=75, max_row=ws.max_row, min_col=1, max_col=1):
        cell = row[0]
        if cell.value:
            task_name = str(cell.value).strip()
            row_idx = cell.row
            if ws.cell(row=row_idx, column=4).__class__.__name__ == 'MergedCell':
                continue
            total_hours = 0
            match_assembly = find_task(task_name, assembly)
            match_testing = find_task(task_name, testing)
            if match_assembly:
                total_hours += match_assembly.get("Hours", 0)
            if match_testing:
                total_hours += match_testing.get("Hours", 0)
            ws.cell(row=row_idx, column=4, value=total_hours)

    wb.save(output_path)
