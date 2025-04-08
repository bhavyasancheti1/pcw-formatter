# === IMPORT REQUIRED LIBRARIES ===
import pandas as pd                    # For reading GPT-generated Excel data
from openpyxl import load_workbook     # For editing Excel templates
from openpyxl.utils import get_column_letter

# === FILE PATHS (CHANGE THESE IF NEEDED) ===
gpt_output_path = 'PCW_MCC_Complete_Parts_and_Labor.xlsx'  # GPT-generated output
pcw_template_path = 'Copy of PCW MULTI with SINGLE INSTALL Jan 2025.xlsx'  # PCW template
output_path = 'Final_PCW_Filled.xlsx'  # Output file name

# === LOAD DATA FROM FILES ===
gpt_data = pd.read_excel(gpt_output_path, sheet_name=None)  # Read all sheets from GPT output
wb = load_workbook(pcw_template_path)  # Load the PCW Excel template
ws = wb['Rev 0']  # Get the worksheet where we’ll paste the data

# === PARTS LIST MAPPING (GPT Parts List → PCW Template) ===
parts_df = gpt_data['Parts List'][['Description', 'Manufacturer', 'Part Number', 'Quantity', 'Cost/ea']].copy()
parts_df = parts_df.fillna('')  # Fill any missing values with empty string
start_row_parts = 10  # Start pasting parts from row 10

# Loop through each row and paste values into PCW template
for idx, row in parts_df.iterrows():
    ws.cell(row=start_row_parts + idx, column=1, value=row['Description'])     # Column A
    ws.cell(row=start_row_parts + idx, column=2, value=row['Manufacturer'])    # Column B
    ws.cell(row=start_row_parts + idx, column=3, value=row['Part Number'])     # Column C
    ws.cell(row=start_row_parts + idx, column=4, value=row['Quantity'])        # Column D
    ws.cell(row=start_row_parts + idx, column=5, value=row['Cost/ea'])         # Column E

# Add Excel formulas to calculate totals
ws.cell(row=39, column=6, value=f'=SUM(F{start_row_parts}:F38)')  # Total Cost (Column F)
ws.cell(row=39, column=8, value=f'=SUM(H{start_row_parts}:H38)')  # Subtotal (Column H)

# === ENGINEERING LABOR MAPPING (GPT Labor Sheet → PCW Rows 43 to 74) ===
eng_df = gpt_data['Engineering Labor'][['Task', '# of DWG / Tasks', 'Hrs per DWG/Task', 'Total Hrs']].copy()
eng_df['Task'] = eng_df['Task'].str.strip()  # Remove leading/trailing spaces from task names

# Loop through PCW template rows 43–74 and match with GPT task names
for row_idx in range(43, 75):
    task_cell = ws.cell(row=row_idx, column=1).value  # Read task name from column A
    if not task_cell:
        continue
    task_clean = task_cell.strip()  # Remove leading/trailing spaces
    match = eng_df[eng_df['Task'] == task_clean]  # Try to find matching task in GPT output
    if not match.empty:
        # Paste values if match is found
        ws.cell(row=row_idx, column=2, value=match.iloc[0]['# of DWG / Tasks'])   # Column B
        ws.cell(row=row_idx, column=3, value=match.iloc[0]['Hrs per DWG/Task'])   # Column C
        ws.cell(row=row_idx, column=4, value=match.iloc[0]['Total Hrs'])          # Column D
    else:
        # If no match found, fill with 0
        ws.cell(row=row_idx, column=2, value=0)
        ws.cell(row=row_idx, column=3, value=0)
        ws.cell(row=row_idx, column=4, value=0)

# === SAVE THE FILLED-IN TEMPLATE ===
wb.save(output_path)
print(f"✅ PCW template filled and saved to: {output_path}")
