import pandas as pd
from openpyxl import load_workbook

# === FILE PATHS ===
gpt_output_path = 'PCW_MCC_Complete_Parts_and_Labor.xlsx'
pcw_template_path = 'Copy of PCW MULTI with SINGLE INSTALL Jan 2025.xlsx'
output_path = 'Final_PCW_Filled_With_All_Labors.xlsx'

# === LOAD FILES ===
gpt_data = pd.read_excel(gpt_output_path, sheet_name=None)
wb = load_workbook(pcw_template_path)
ws = wb['Rev 0']

# === PARTS LIST MAPPING ===
parts_df = gpt_data['Parts List'][['Description', 'Manufacturer', 'Part Number', 'Quantity', 'Cost/ea']].copy()
parts_df = parts_df.fillna('')
start_row_parts = 10

for idx, row in parts_df.iterrows():
    ws.cell(row=start_row_parts + idx, column=1, value=row['Description'])
    ws.cell(row=start_row_parts + idx, column=2, value=row['Manufacturer'])
    ws.cell(row=start_row_parts + idx, column=3, value=row['Part Number'])
    ws.cell(row=start_row_parts + idx, column=4, value=row['Quantity'])
    ws.cell(row=start_row_parts + idx, column=5, value=row['Cost/ea'])

ws.cell(row=39, column=6, value=f'=SUM(F{start_row_parts}:F38)')  # Total Cost
ws.cell(row=39, column=8, value=f'=SUM(H{start_row_parts}:H38)')  # Subtotal

# === ENGINEERING LABOR MAPPING ===
eng_df = gpt_data['Engineering Labor'][['Task', '# of DWG / Tasks', 'Hrs per DWG/Task', 'Total Hrs']].copy()
eng_df['Task'] = eng_df['Task'].str.strip()

for row_idx in range(43, 75):  # Rows 43 to 74
    task_cell = ws.cell(row=row_idx, column=1).value
    if not task_cell:
        continue
    task_clean = task_cell.strip()
    match = eng_df[eng_df['Task'] == task_clean]
    if not match.empty:
        ws.cell(row=row_idx, column=2, value=match.iloc[0]['# of DWG / Tasks'])
        ws.cell(row=row_idx, column=3, value=match.iloc[0]['Hrs per DWG/Task'])
        ws.cell(row=row_idx, column=4, value=match.iloc[0]['Total Hrs'])
    else:
        ws.cell(row=row_idx, column=2, value=0)
        ws.cell(row=row_idx, column=3, value=0)
        ws.cell(row=row_idx, column=4, value=0)



# === ASSEMBLY + IN-HOUSE TESTING LABOR MAPPING ===
# Combine hours and map to Column D (Total Hours)

assembly_df = gpt_data['Assembly Labor'][['Task', 'Hours']].copy()
assembly_df['Task'] = assembly_df['Task'].str.strip()

testing_df = gpt_data['In-House Testing Labor'][['Task', 'Hours']].copy()
testing_df['Task'] = testing_df['Task'].str.strip()

# Go through each task row in PCW template starting from row 75
for row in ws.iter_rows(min_row=75, max_row=ws.max_row, min_col=1, max_col=1):
    cell = row[0]
    if cell.value:
        task_name = str(cell.value).strip()
        row_idx = cell.row

        # Skip merged cells to avoid openpyxl error
        if ws.cell(row=row_idx, column=4).__class__.__name__ == 'MergedCell':
            continue

        # Combine hours from Assembly and In-House Testing
        total_hours = 0
        match_assembly = assembly_df[assembly_df['Task'] == task_name]
        match_testing = testing_df[testing_df['Task'] == task_name]

        if not match_assembly.empty:
            total_hours += match_assembly.iloc[0]['Hours']
        if not match_testing.empty:
            total_hours += match_testing.iloc[0]['Hours']

        #  Write to Column D (Total Hours)
        ws.cell(row=row_idx, column=4, value=total_hours)



# === SAVE FINAL FILE ===
wb.save(output_path)
print(f"PCW template fully filled and saved to: {output_path}")
